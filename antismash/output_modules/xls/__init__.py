# vim: set fileencoding=utf-8 :
#
# Copyright (C) 2010-2012 Marnix H. Medema
# University of Groningen
# Department of Microbial Physiology / Groningen Bioinformatics Centre
#
# Copyright (C) 2011,2012 Kai Blin
# University of Tuebingen
# Interfaculty Institute of Microbiology and Infection Medicine
# Div. of Microbiology/Biotechnology
#
# Copyright (C) 2025 Elena Del Pup
# Wageningen University 
# Department of Plant Sciences 
# Bioinformatics group  
#
# License: GNU Affero General Public License v3 or later
# A copy of GNU AGPL v3 should have been included in this software package in LICENSE.txt.

"""Excel and TXT output format module

"""

from openpyxl import Workbook
from openpyxl.styles import Font
import os
from antismash import utils
from os import path

name = "xls"
short_description = "xls_txt_output"
priority = 1

def write(seq_records, options, filename="plantgeneclusters.txt", directory=None):
    """
    Write gene cluster data to TXT and XLS files.

    Parameters:
    - seq_records: List of SeqRecord objects to process.
    - options: The options object containing configuration settings.
    - filename: (Optional) Name of the output TXT file. Defaults to 'plantgeneclusters.txt'.
    - directory: (Optional) Directory to write the output files. Defaults to 'options.full_outputfolder_path'.
    """
    # Determine output file name and directory
    outfolder = options.full_outputfolder_path
    directory = directory or outfolder

    # Skip processing for protein input type
    if options.input_type == 'prot':
        return

    # Open TXT file for writing
    txtfile_path = path.join(directory, filename)
    with open(txtfile_path, "w") as txtfile:
        # Initialize Excel Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Gene Clusters"

        # Define header font
        header_font = Font(bold=True)

        # Write Excel headers
        headers = [
            "Input accession number", "Input name", "Gene cluster type",
            "Gene cluster genes", "Gene cluster gene accessions"
        ]
        if options.knownclusterblast:
            headers.append("Compound with gene cluster of highest homology")

        for col_num, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col_num, value=header).font = header_font

        # Write gene cluster data
        row = 2
        for seq_record in seq_records:
            clusters = utils.get_sorted_cluster_features(seq_record)
            for cluster in clusters:
                clustertype = utils.get_cluster_type(cluster)
                clusternr = utils.get_cluster_number(cluster)
                clustergenes = [
                    utils.get_gene_id(cds) for cds in utils.get_cluster_cds_features(cluster, seq_record)
                ]
                accessions = [
                    utils.get_gene_acc(cds) for cds in utils.get_cluster_cds_features(cluster, seq_record)
                ]

                # Write data to Excel
                ws.cell(row=row, column=1, value=seq_record.id)
                ws.cell(row=row, column=2, value=seq_record.description[:32767])  # Excel cell limit
                ws.cell(row=row, column=3, value=clustertype)
                ws.cell(row=row, column=4, value=";".join(clustergenes[:32767]))
                ws.cell(row=row, column=5, value=";".join(accessions[:32767]))

                if hasattr(seq_record, 'closestcompounddict') and \
                   clusternr in seq_record.closestcompounddict:
                    ws.cell(row=row, column=6, value=seq_record.closestcompounddict[clusternr])

                # Write data to TXT file
                txtfile.write("\t".join([
                    seq_record.id, seq_record.description,
                    "c" + str(utils.get_cluster_number(cluster)),
                    clustertype, ";".join(clustergenes), ";".join(accessions)
                ]) + "\n")

                row += 1

        # Save the Excel workbook
        excel_file_path = path.join(outfolder, f"{seq_record.id}.geneclusters.xlsx")
        wb.save(excel_file_path)
        print(f"Saved Excel file: {excel_file_path}")
